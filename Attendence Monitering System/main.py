import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timedelta


class AttendanceManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook, self.sheet = self.load_workbook()
        self.attendance_data = {}

    def load_workbook(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
        except FileNotFoundError:
            print("File not found. Creating a new workbook.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Attendance"
            sheet.cell(row=1, column=1).value = "Student Name"
            sheet.cell(row=1, column=2).value = "Roll Number"
        return workbook, sheet

    def save_workbook(self):
        self.calculate_detained()
        self.workbook.save(self.file_path)

    def mark_attendance(self, student_name, roll_number, attendance_date, status):
        if roll_number not in self.attendance_data:
            self.attendance_data[roll_number] = {'name': student_name, 'total_days': 0, 'present_days': 0}
        self.attendance_data[roll_number]['total_days'] += 1
        if status == 'P':
            self.attendance_data[roll_number]['present_days'] += 1

        student_row = None
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == student_name and self.sheet.cell(row=row,
                                                                                            column=2).value == roll_number:
                student_row = row
                break

        if student_row is None:
            id_col = 2
            id_col_letter = get_column_letter(id_col)
            self.sheet.column_dimensions[id_col_letter].width = 20
            student_col = 1
            student_col_letter = get_column_letter(student_col)
            self.sheet.column_dimensions[student_col_letter].width = 25
            student_row = self.sheet.max_row + 1
            self.sheet.cell(row=student_row, column=1).value = student_name
            self.sheet.cell(row=student_row, column=2).value = roll_number

        date_col = None
        for col in range(3, self.sheet.max_column + 1):
            if self.sheet.cell(row=1, column=col).value == attendance_date:
                date_col = col
                break

        if date_col is None:
            date_col = self.sheet.max_column + 1
            self.sheet.cell(row=1, column=date_col).value = attendance_date
            col_letter = get_column_letter(date_col)
            self.sheet.column_dimensions[col_letter].width = 15
            self.sheet.cell(row=1, column=date_col).alignment = Alignment(horizontal="center")

        self.sheet.cell(row=student_row, column=date_col).value = status

    def calculate_detained(self):
        detained_col = None
        for col in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(row=1, column=col).value == "Detained":
                detained_col = col
                break

        if detained_col is None:
            detained_col = self.sheet.max_column + 1
            self.sheet.cell(row=1, column=detained_col).value = "Detained"
            col_letter = get_column_letter(detained_col)
            self.sheet.column_dimensions[col_letter].width = 15
            self.sheet.cell(row=1, column=detained_col).alignment = Alignment(horizontal="center")

        for row in range(2, self.sheet.max_row + 1):
            roll_number = self.sheet.cell(row=row, column=2).value
            if roll_number in self.attendance_data:
                student_data = self.attendance_data[roll_number]
                total_days = student_data['total_days']
                present_days = student_data['present_days']
                if total_days > 0:
                    attendance_percentage = (present_days / total_days) * 100
                    detained_status = "Yes" if attendance_percentage < 75 else "No"
                    self.sheet.cell(row=row, column=detained_col).value = detained_status


def get_thursdays_within_date_range(start_date, end_date, holidays):
    thursdays = []
    current_date = start_date

    while current_date <= end_date:
        if current_date.weekday() == 3 and current_date not in holidays:
            thursdays.append(current_date)
        current_date += timedelta(days=1)

    return thursdays


def main():
    file_path = "attendance.xlsx"
    attendance_manager = AttendanceManager(file_path)

    students = [
        {"name": "Kaushik Bagde", "roll_number": "23005019"},
        {"name": "Krushna Mandare", "roll_number": "23005021"},
        {"name": "Madhav Emineni", "roll_number": "23005023"},
        {"name": "Milind Kumar", "roll_number": "23005024"},
        {"name": "Mithilesh Potbhare", "roll_number": "23005025"},
        {"name": "Nikhil Parashar", "roll_number": "23005026"},
        {"name": "Nikita Rathod", "roll_number": "23005027"},
        {"name": "Nishad Bhale", "roll_number": "23005028"}
    ]

    start_date = datetime.strptime("2024-07-01", "%Y-%m-%d")
    end_date = datetime.strptime("2024-10-31", "%Y-%m-%d")

    holidays = [
        datetime.strptime("2024-07-04", "%Y-%m-%d"),
        datetime.strptime("2024-08-15", "%Y-%m-%d"),
        datetime.strptime("2024-10-31", "%Y-%m-%d"),
        # datetime.strptime("2024-09-", "%Y-%m-%d")
    ]

    thursdays = get_thursdays_within_date_range(start_date, end_date, holidays)

    while True:
        date_str = input("Enter the date (YYYY-MM-DD) to mark attendance for Batch 2 or 'exit' to quit: ").strip()
        if date_str.lower() == 'exit':
            break

        try:
            date_input = datetime.strptime(date_str, "%Y-%m-%d")
            if date_input in thursdays:
                print(f"Marking attendance for {date_str} : Batch 2")
                for student in students:
                    status = input(
                        f"Mark attendance for {student['name']} (Roll Number: {student['roll_number']}) on {date_str} [P/A]: ").strip().upper()
                    while status not in ['P', 'A']:
                        print("Invalid input. Please enter 'P' for Present or 'A' for Absent.")
                        status = input(
                            f"Mark attendance for {student['name']} (Roll Number: {student['roll_number']}) on {date_str} [P/A]: ").strip().upper()
                    attendance_manager.mark_attendance(student["name"], student["roll_number"], date_str, status)
            else:
                print(f"{date_str} is not a valid Thursday within the specified date range or is a holiday.")
        except ValueError:
            print("Invalid date format. Please enter the date in YYYY-MM-DD format.")

    attendance_manager.save_workbook()
    print("Attendance updated successfully.")


if __name__ == "__main__":
    main()
